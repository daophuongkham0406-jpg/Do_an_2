from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

# ============================================================
# Workout History Validator v1.0
# Validates workout_history_master.xlsx and, when supplied,
# cross-checks User / Workout Plan / Exercise master workbooks.
# ============================================================

SESSION_SHEET = "Workout_History_Sessions"
ITEM_SHEET = "Workout_History_Items"
SUMMARY_SHEET = "Workout_History_Summary"

SESSION_REQUIRED_COLUMNS = {
    "history_session_id", "user_id", "plan_id", "week_number", "day_number",
    "planned_day_name", "planned_session_name", "scheduled_date", "completion_status",
    "planned_item_count", "completed_item_count", "completion_pct",
    "planned_working_sets", "completed_working_sets", "set_completion_pct",
    "session_duration_target_min", "actual_duration_min", "session_rpe",
    "energy_before", "fatigue_after", "sleep_hours_snapshot",
    "body_weight_kg_snapshot", "pain_reported", "pain_areas", "readiness_score",
    "recovery_flag", "record_source", "is_synthetic", "notes", "created_at",
}

ITEM_REQUIRED_COLUMNS = {
    "history_item_id", "history_session_id", "user_id", "plan_id", "plan_item_id",
    "exercise_id", "exercise_name_snapshot", "exercise_order", "planned_sets",
    "planned_rep_min", "planned_rep_max", "planned_target_rpe", "planned_rest_seconds",
    "actual_sets_completed", "actual_reps_json", "actual_load_kg", "actual_rpe",
    "completion_status", "pain_during_exercise", "pain_areas", "technique_quality",
    "difficulty_rating", "exercise_enjoyment", "feedback_signal", "record_source",
    "is_synthetic", "notes", "created_at",
}

SUMMARY_REQUIRED_COLUMNS = {
    "summary_id", "user_id", "plan_id", "representative_week", "representative_day",
    "session_status", "session_completion_pct", "set_completion_pct", "session_rpe",
    "fatigue_after", "pain_reported", "avg_difficulty", "avg_enjoyment",
    "positive_items", "neutral_items", "negative_items", "recovery_flag",
    "progression_recommendation",
}

SESSION_STATUS = {"Completed", "Partial", "Skipped"}
ITEM_STATUS = {"Completed", "Modified", "Skipped"}
YES_NO = {"Yes", "No"}
RECOVERY_FLAGS = {"Ready", "Monitor", "Review"}
FEEDBACK = {"Positive", "Neutral", "Negative"}
TECHNIQUE = {"Good", "Fair", "Poor"}
RECORD_SOURCES = {"Synthetic", "App", "Manual", "Imported"}
PROGRESSION = {
    "PROGRESS_IF_RECOVERED", "MAINTAIN", "HOLD_AND_REVIEW_ADHERENCE",
    "HOLD_OR_REDUCE_DEMAND", "REVIEW_BEFORE_PROGRESSION",
}
WEEKDAYS = {"Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"}

ID_PATTERNS = {
    "history_session_id": re.compile(r"^WHS\d{8}$"),
    "history_item_id": re.compile(r"^WHI\d{9}$"),
    "summary_id": re.compile(r"^WHSUM\d{6}$"),
    "user_id": re.compile(r"^U\d{6}$"),
    "plan_id": re.compile(r"^PLAN\d{6}$"),
    "plan_item_id": re.compile(r"^WPI\d{8}$"),
    "exercise_id": re.compile(r"^EX\d{4}$"),
}


@dataclass
class Issue:
    severity: str
    rule_id: str
    sheet: str
    row: int | None
    key: str
    field: str
    message: str
    value: Any = None

    def render(self) -> str:
        loc = self.sheet
        if self.row is not None:
            loc += f" | dòng {self.row}"
        if self.key:
            loc += f" | {self.key}"
        if self.field:
            loc += f" | {self.field}"
        text = f"[{self.severity}] {self.rule_id} ({loc}): {self.message}"
        if self.value not in (None, ""):
            text += f"\n    Giá trị: {self.value}"
        return text


class Validator:
    def __init__(self) -> None:
        self.issues: list[Issue] = []

    def add(self, severity: str, rule_id: str, sheet: str, row: int | None,
            key: str, field: str, message: str, value: Any = None) -> None:
        self.issues.append(Issue(severity, rule_id, sheet, row, key, field, message, value))

    def error(self, *args, **kwargs):
        self.add("ERROR", *args, **kwargs)

    def warning(self, *args, **kwargs):
        self.add("WARNING", *args, **kwargs)

    def info(self, *args, **kwargs):
        self.add("INFO", *args, **kwargs)


def is_blank(v: Any) -> bool:
    return v is None or (isinstance(v, str) and not v.strip())


def as_float(v: Any) -> float | None:
    if is_blank(v):
        return None
    try:
        x = float(v)
        return x if math.isfinite(x) else None
    except (TypeError, ValueError):
        return None


def as_int(v: Any) -> int | None:
    x = as_float(v)
    if x is None or not float(x).is_integer():
        return None
    return int(x)


def parse_json_list(v: Any) -> tuple[list[Any] | None, str | None]:
    if isinstance(v, list):
        return v, None
    if is_blank(v):
        return [], None
    try:
        data = json.loads(str(v))
    except Exception as exc:
        return None, str(exc)
    if not isinstance(data, list):
        return None, "JSON value is not an array"
    return data, None


def pct(numerator: float, denominator: float) -> float:
    return round(100.0 * numerator / denominator, 1) if denominator else 0.0


def approx_equal(a: Any, b: Any, tol: float = 0.11) -> bool:
    x, y = as_float(a), as_float(b)
    return x is not None and y is not None and abs(x - y) <= tol


def load_rows(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        raise KeyError(f"Không tìm thấy sheet '{sheet_name}' trong {path.name}")
    ws = wb[sheet_name]
    iterator = ws.iter_rows(values_only=True)
    try:
        headers = [str(x).strip() if x is not None else "" for x in next(iterator)]
    except StopIteration:
        wb.close()
        return []
    rows = []
    for excel_row, values in enumerate(iterator, start=2):
        if not any(not is_blank(v) for v in values):
            continue
        row = dict(zip(headers, values))
        row["__row__"] = excel_row
        rows.append(row)
    wb.close()
    return rows


def get_headers(path: Path, sheet_name: str) -> set[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return set()
    ws = wb[sheet_name]
    values = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
    wb.close()
    return {str(v).strip() for v in values if v is not None}


def validate_schema(v: Validator, history_path: Path) -> bool:
    wb = load_workbook(history_path, read_only=True, data_only=True)
    sheets = set(wb.sheetnames)
    wb.close()
    required_sheets = {SESSION_SHEET, ITEM_SHEET, SUMMARY_SHEET}
    missing_sheets = required_sheets - sheets
    for s in sorted(missing_sheets):
        v.error("HIS001", "Workbook", None, "", "sheet", f"Thiếu sheet bắt buộc: {s}")
    if missing_sheets:
        return False

    for sheet, required in [
        (SESSION_SHEET, SESSION_REQUIRED_COLUMNS),
        (ITEM_SHEET, ITEM_REQUIRED_COLUMNS),
        (SUMMARY_SHEET, SUMMARY_REQUIRED_COLUMNS),
    ]:
        headers = get_headers(history_path, sheet)
        for col in sorted(required - headers):
            v.error("HIS002", sheet, 1, "", col, "Thiếu cột bắt buộc")
    return not any(i.severity == "ERROR" and i.rule_id in {"HIS001", "HIS002"} for i in v.issues)


def check_unique_and_id_format(v: Validator, rows: list[dict], sheet: str,
                               id_field: str, rule_prefix: str) -> None:
    seen: dict[str, int] = {}
    pattern = ID_PATTERNS[id_field]
    for r in rows:
        row_no = r["__row__"]
        raw = r.get(id_field)
        key = str(raw or "")
        if is_blank(raw):
            v.error(f"{rule_prefix}01", sheet, row_no, key, id_field, "ID bắt buộc không được rỗng")
            continue
        if not pattern.match(key):
            v.error(f"{rule_prefix}02", sheet, row_no, key, id_field, "Sai định dạng ID", raw)
        if key in seen:
            v.error(f"{rule_prefix}03", sheet, row_no, key, id_field,
                    f"ID bị trùng; đã xuất hiện ở dòng {seen[key]}", raw)
        else:
            seen[key] = row_no


def validate_sessions(v: Validator, rows: list[dict]) -> None:
    check_unique_and_id_format(v, rows, SESSION_SHEET, "history_session_id", "SES")
    for r in rows:
        n = r["__row__"]
        sid = str(r.get("history_session_id") or "")

        for field in ["user_id", "plan_id"]:
            raw = r.get(field)
            if is_blank(raw):
                v.error("SES004", SESSION_SHEET, n, sid, field, "Foreign key bắt buộc không được rỗng")
            elif not ID_PATTERNS[field].match(str(raw)):
                v.error("SES005", SESSION_SHEET, n, sid, field, "Sai định dạng foreign key", raw)

        week = as_int(r.get("week_number"))
        day = as_int(r.get("day_number"))
        if week is None or not 1 <= week <= 52:
            v.error("SES006", SESSION_SHEET, n, sid, "week_number", "week_number phải là số nguyên 1..52", r.get("week_number"))
        if day is None or not 1 <= day <= 7:
            v.error("SES007", SESSION_SHEET, n, sid, "day_number", "day_number phải là số nguyên 1..7", r.get("day_number"))
        if r.get("planned_day_name") not in WEEKDAYS:
            v.error("SES008", SESSION_SHEET, n, sid, "planned_day_name", "Tên ngày không hợp lệ", r.get("planned_day_name"))

        status = r.get("completion_status")
        if status not in SESSION_STATUS:
            v.error("SES009", SESSION_SHEET, n, sid, "completion_status", f"Chỉ nhận {sorted(SESSION_STATUS)}", status)

        pi = as_int(r.get("planned_item_count"))
        ci = as_int(r.get("completed_item_count"))
        ps = as_int(r.get("planned_working_sets"))
        cs = as_int(r.get("completed_working_sets"))
        if pi is None or pi < 0:
            v.error("SES010", SESSION_SHEET, n, sid, "planned_item_count", "Phải là số nguyên >= 0", r.get("planned_item_count"))
        if ci is None or ci < 0 or (pi is not None and ci > pi):
            v.error("SES011", SESSION_SHEET, n, sid, "completed_item_count", "Phải thỏa 0 <= completed <= planned", r.get("completed_item_count"))
        if ps is None or ps < 0:
            v.error("SES012", SESSION_SHEET, n, sid, "planned_working_sets", "Phải là số nguyên >= 0", r.get("planned_working_sets"))
        if cs is None or cs < 0 or (ps is not None and cs > ps):
            v.error("SES013", SESSION_SHEET, n, sid, "completed_working_sets", "Phải thỏa 0 <= completed <= planned", r.get("completed_working_sets"))

        for field in ["completion_pct", "set_completion_pct"]:
            x = as_float(r.get(field))
            if x is None or not 0 <= x <= 100:
                v.error("SES014", SESSION_SHEET, n, sid, field, "Phần trăm phải nằm trong 0..100", r.get(field))

        if pi is not None and ci is not None and not approx_equal(r.get("completion_pct"), pct(ci, pi)):
            v.error("SES015", SESSION_SHEET, n, sid, "completion_pct", "Không khớp completed_item_count/planned_item_count", r.get("completion_pct"))
        if ps is not None and cs is not None and not approx_equal(r.get("set_completion_pct"), pct(cs, ps)):
            v.error("SES016", SESSION_SHEET, n, sid, "set_completion_pct", "Không khớp completed_working_sets/planned_working_sets", r.get("set_completion_pct"))

        duration_target = as_float(r.get("session_duration_target_min"))
        duration_actual = as_float(r.get("actual_duration_min"))
        if duration_target is None or duration_target <= 0 or duration_target > 240:
            v.error("SES017", SESSION_SHEET, n, sid, "session_duration_target_min", "Thời lượng mục tiêu phải >0 và <=240 phút", r.get("session_duration_target_min"))
        if duration_actual is None or duration_actual < 0 or duration_actual > 360:
            v.error("SES018", SESSION_SHEET, n, sid, "actual_duration_min", "Thời lượng thực tế phải 0..360 phút", r.get("actual_duration_min"))

        if status == "Skipped":
            if (ci or 0) != 0 or (cs or 0) != 0 or (duration_actual or 0) != 0:
                v.error("SES019", SESSION_SHEET, n, sid, "completion_status", "Skipped phải có completed_items=0, completed_sets=0, duration=0")
            if not is_blank(r.get("session_rpe")):
                v.error("SES020", SESSION_SHEET, n, sid, "session_rpe", "Skipped không được có session_rpe", r.get("session_rpe"))
        elif status in {"Completed", "Partial"}:
            if (ci or 0) <= 0 or (cs or 0) <= 0:
                v.error("SES021", SESSION_SHEET, n, sid, "completion_status", "Completed/Partial phải có actual work > 0")

        rpe = as_float(r.get("session_rpe"))
        if not is_blank(r.get("session_rpe")) and (rpe is None or not 1 <= rpe <= 10):
            v.error("SES022", SESSION_SHEET, n, sid, "session_rpe", "RPE phải trong 1..10 hoặc để trống khi Skipped", r.get("session_rpe"))
        if rpe is not None and rpe >= 9.5:
            v.warning("HIS008", SESSION_SHEET, n, sid, "session_rpe", "RPE >= 9.5: effort rất cao, nên xem xét tải/recovery", rpe)

        for field in ["energy_before", "fatigue_after"]:
            if is_blank(r.get(field)) and field == "fatigue_after" and status == "Skipped":
                continue
            x = as_int(r.get(field))
            if x is None or not 1 <= x <= 5:
                v.error("SES023", SESSION_SHEET, n, sid, field, "Phải là số nguyên 1..5", r.get(field))

        sleep = as_float(r.get("sleep_hours_snapshot"))
        if sleep is None or not 0 <= sleep <= 24:
            v.error("SES024", SESSION_SHEET, n, sid, "sleep_hours_snapshot", "Giờ ngủ phải trong 0..24", r.get("sleep_hours_snapshot"))
        elif sleep < 5:
            v.warning("SES025", SESSION_SHEET, n, sid, "sleep_hours_snapshot", "Giấc ngủ rất thấp; recovery có thể bị ảnh hưởng", sleep)

        bw = as_float(r.get("body_weight_kg_snapshot"))
        if bw is None or not 20 <= bw <= 400:
            v.warning("SES026", SESSION_SHEET, n, sid, "body_weight_kg_snapshot", "Cân nặng ngoài khoảng kiểm tra hợp lý 20..400 kg", r.get("body_weight_kg_snapshot"))

        pain = r.get("pain_reported")
        pain_areas, err = parse_json_list(r.get("pain_areas"))
        if pain not in YES_NO:
            v.error("SES027", SESSION_SHEET, n, sid, "pain_reported", "Chỉ nhận Yes/No", pain)
        if err:
            v.error("SES028", SESSION_SHEET, n, sid, "pain_areas", "pain_areas phải là JSON array hợp lệ", r.get("pain_areas"))
        elif pain == "Yes" and not pain_areas:
            v.error("SES029", SESSION_SHEET, n, sid, "pain_areas", "pain_reported=Yes thì pain_areas không được rỗng")
        elif pain == "No" and pain_areas:
            v.warning("SES030", SESSION_SHEET, n, sid, "pain_areas", "pain_reported=No nhưng pain_areas lại có dữ liệu", pain_areas)

        readiness = as_float(r.get("readiness_score"))
        if readiness is None or not 1 <= readiness <= 10:
            v.error("SES031", SESSION_SHEET, n, sid, "readiness_score", "Readiness phải trong 1..10", r.get("readiness_score"))

        recovery = r.get("recovery_flag")
        if recovery not in RECOVERY_FLAGS:
            v.error("SES032", SESSION_SHEET, n, sid, "recovery_flag", f"Chỉ nhận {sorted(RECOVERY_FLAGS)}", recovery)
        if pain == "Yes" and recovery != "Review":
            v.warning("HIS006", SESSION_SHEET, n, sid, "recovery_flag", "Có pain nhưng recovery_flag không phải Review", recovery)
        if r.get("completion_status") != "Skipped" and as_float(r.get("set_completion_pct")) is not None and float(r["set_completion_pct"]) < 70:
            v.warning("HIS007", SESSION_SHEET, n, sid, "set_completion_pct", "Set completion < 70%; cần xem lại adherence hoặc demand", r.get("set_completion_pct"))

        source = r.get("record_source")
        if source not in RECORD_SOURCES:
            v.warning("SES033", SESSION_SHEET, n, sid, "record_source", f"Nguồn ngoài taxonomy {sorted(RECORD_SOURCES)}", source)
        synthetic = r.get("is_synthetic")
        if source == "Synthetic" and synthetic not in (True, 1, "TRUE", "True", "true"):
            v.error("HIS010", SESSION_SHEET, n, sid, "is_synthetic", "record_source=Synthetic thì is_synthetic phải True", synthetic)


def validate_items(v: Validator, rows: list[dict]) -> None:
    check_unique_and_id_format(v, rows, ITEM_SHEET, "history_item_id", "ITM")
    for r in rows:
        n = r["__row__"]
        iid = str(r.get("history_item_id") or "")
        for field in ["history_session_id", "user_id", "plan_id", "plan_item_id", "exercise_id"]:
            raw = r.get(field)
            if is_blank(raw):
                v.error("ITM004", ITEM_SHEET, n, iid, field, "Foreign key bắt buộc không được rỗng")
            elif field in ID_PATTERNS and not ID_PATTERNS[field].match(str(raw)):
                v.error("ITM005", ITEM_SHEET, n, iid, field, "Sai định dạng ID", raw)

        order = as_int(r.get("exercise_order"))
        if order is None or order < 1:
            v.error("ITM006", ITEM_SHEET, n, iid, "exercise_order", "exercise_order phải là số nguyên >= 1", r.get("exercise_order"))

        psets = as_int(r.get("planned_sets"))
        aset = as_int(r.get("actual_sets_completed"))
        if psets is None or psets < 0:
            v.error("ITM007", ITEM_SHEET, n, iid, "planned_sets", "planned_sets phải là số nguyên >= 0", r.get("planned_sets"))
        if aset is None or aset < 0 or (psets is not None and aset > psets):
            v.error("HIS003", ITEM_SHEET, n, iid, "actual_sets_completed", "Phải thỏa 0 <= actual_sets_completed <= planned_sets", r.get("actual_sets_completed"))

        rmin, rmax = as_int(r.get("planned_rep_min")), as_int(r.get("planned_rep_max"))
        if rmin is not None and rmax is not None and (rmin <= 0 or rmax <= 0 or rmin > rmax):
            v.error("ITM008", ITEM_SHEET, n, iid, "planned_rep_min/planned_rep_max", "Rep range phải dương và min <= max", f"{rmin}..{rmax}")

        target = as_float(r.get("planned_target_rpe"))
        if target is None or not 1 <= target <= 10:
            v.error("ITM009", ITEM_SHEET, n, iid, "planned_target_rpe", "Target RPE phải trong 1..10", r.get("planned_target_rpe"))
        rest = as_int(r.get("planned_rest_seconds"))
        if rest is None or not 0 <= rest <= 900:
            v.error("ITM010", ITEM_SHEET, n, iid, "planned_rest_seconds", "Rest phải là số nguyên 0..900 giây", r.get("planned_rest_seconds"))

        reps, err = parse_json_list(r.get("actual_reps_json"))
        if err:
            v.error("ITM011", ITEM_SHEET, n, iid, "actual_reps_json", "Phải là JSON array hợp lệ", r.get("actual_reps_json"))
        elif aset is not None and len(reps) != aset:
            v.error("ITM012", ITEM_SHEET, n, iid, "actual_reps_json", "Số phần tử reps phải bằng actual_sets_completed", f"sets={aset}, reps={reps}")
        elif any(as_int(x) is None or as_int(x) <= 0 for x in reps):
            v.error("ITM013", ITEM_SHEET, n, iid, "actual_reps_json", "Mỗi rep phải là số nguyên > 0", reps)

        load = as_float(r.get("actual_load_kg"))
        if not is_blank(r.get("actual_load_kg")) and (load is None or load < 0 or load > 1000):
            v.error("ITM014", ITEM_SHEET, n, iid, "actual_load_kg", "Load phải trong 0..1000 kg hoặc để trống", r.get("actual_load_kg"))

        arpe = as_float(r.get("actual_rpe"))
        if not is_blank(r.get("actual_rpe")) and (arpe is None or not 1 <= arpe <= 10):
            v.error("ITM015", ITEM_SHEET, n, iid, "actual_rpe", "Actual RPE phải trong 1..10 hoặc để trống khi Skipped", r.get("actual_rpe"))

        status = r.get("completion_status")
        if status not in ITEM_STATUS:
            v.error("ITM016", ITEM_SHEET, n, iid, "completion_status", f"Chỉ nhận {sorted(ITEM_STATUS)}", status)
        if status == "Skipped":
            if (aset or 0) != 0:
                v.error("HIS004", ITEM_SHEET, n, iid, "actual_sets_completed", "Skipped => actual_sets_completed phải bằng 0", aset)
            if reps:
                v.error("ITM017", ITEM_SHEET, n, iid, "actual_reps_json", "Skipped => actual_reps_json phải rỗng", reps)
            if not is_blank(r.get("actual_rpe")):
                v.error("ITM018", ITEM_SHEET, n, iid, "actual_rpe", "Skipped => actual_rpe phải rỗng", r.get("actual_rpe"))
        elif (aset or 0) <= 0:
            v.error("ITM019", ITEM_SHEET, n, iid, "actual_sets_completed", "Completed/Modified phải có actual_sets_completed > 0", aset)

        pain = r.get("pain_during_exercise")
        pain_areas, err = parse_json_list(r.get("pain_areas"))
        if pain not in YES_NO:
            v.error("ITM020", ITEM_SHEET, n, iid, "pain_during_exercise", "Chỉ nhận Yes/No", pain)
        if err:
            v.error("ITM021", ITEM_SHEET, n, iid, "pain_areas", "pain_areas phải là JSON array", r.get("pain_areas"))
        elif pain == "Yes" and not pain_areas:
            v.error("HIS005", ITEM_SHEET, n, iid, "pain_areas", "pain=Yes thì pain_areas không được rỗng")
        elif pain == "No" and pain_areas:
            v.warning("ITM022", ITEM_SHEET, n, iid, "pain_areas", "pain=No nhưng pain_areas có dữ liệu", pain_areas)
        if pain == "Yes" and status == "Completed":
            v.warning("ITM023", ITEM_SHEET, n, iid, "completion_status", "Có pain nhưng item vẫn Completed; cân nhắc Modified")

        technique = r.get("technique_quality")
        if status == "Skipped":
            if not is_blank(technique):
                v.warning("ITM024", ITEM_SHEET, n, iid, "technique_quality", "Skipped thường không nên có technique_quality", technique)
        elif technique not in TECHNIQUE:
            v.error("ITM025", ITEM_SHEET, n, iid, "technique_quality", f"Chỉ nhận {sorted(TECHNIQUE)}", technique)

        for field in ["difficulty_rating", "exercise_enjoyment"]:
            if status == "Skipped" and is_blank(r.get(field)):
                continue
            x = as_int(r.get(field))
            if x is None or not 1 <= x <= 5:
                v.error("ITM026", ITEM_SHEET, n, iid, field, "Phải là số nguyên 1..5 khi exercise được thực hiện", r.get(field))

        feedback = r.get("feedback_signal")
        if feedback not in FEEDBACK:
            v.error("ITM027", ITEM_SHEET, n, iid, "feedback_signal", f"Chỉ nhận {sorted(FEEDBACK)}", feedback)
        if status == "Skipped" and feedback != "Negative":
            v.warning("ITM028", ITEM_SHEET, n, iid, "feedback_signal", "Skipped nên tạo Negative feedback signal", feedback)
        if pain == "Yes" and feedback != "Negative":
            v.warning("ITM029", ITEM_SHEET, n, iid, "feedback_signal", "Pain nên tạo Negative feedback signal", feedback)

        source = r.get("record_source")
        synthetic = r.get("is_synthetic")
        if source == "Synthetic" and synthetic not in (True, 1, "TRUE", "True", "true"):
            v.error("HIS010", ITEM_SHEET, n, iid, "is_synthetic", "record_source=Synthetic thì is_synthetic phải True", synthetic)
        if source == "Synthetic" and is_blank(r.get("actual_load_kg")):
            v.info("HIS009", ITEM_SHEET, n, iid, "actual_load_kg", "Synthetic history được phép để trống actual_load_kg vì nguồn plan không có kg đáng tin cậy")


def validate_internal_links(v: Validator, sessions: list[dict], items: list[dict], summaries: list[dict]) -> None:
    session_by_id = {str(r.get("history_session_id")): r for r in sessions}
    items_by_session: dict[str, list[dict]] = defaultdict(list)

    for r in items:
        n, iid = r["__row__"], str(r.get("history_item_id") or "")
        sid = str(r.get("history_session_id") or "")
        if sid not in session_by_id:
            v.error("REL001", ITEM_SHEET, n, iid, "history_session_id", "Không tồn tại trong Workout_History_Sessions", sid)
            continue
        s = session_by_id[sid]
        if r.get("user_id") != s.get("user_id"):
            v.error("REL002", ITEM_SHEET, n, iid, "user_id", "user_id không khớp session", r.get("user_id"))
        if r.get("plan_id") != s.get("plan_id"):
            v.error("REL003", ITEM_SHEET, n, iid, "plan_id", "plan_id không khớp session", r.get("plan_id"))
        items_by_session[sid].append(r)

    # Recompute session aggregates from item rows.
    for s in sessions:
        n, sid = s["__row__"], str(s.get("history_session_id") or "")
        group = items_by_session.get(sid, [])
        planned_items = len(group)
        completed_items = sum(r.get("completion_status") != "Skipped" for r in group)
        planned_sets = sum(as_int(r.get("planned_sets")) or 0 for r in group)
        completed_sets = sum(as_int(r.get("actual_sets_completed")) or 0 for r in group)
        pain_yes = any(r.get("pain_during_exercise") == "Yes" for r in group)
        rpes = [as_float(r.get("actual_rpe")) for r in group if as_float(r.get("actual_rpe")) is not None]
        mean_rpe = round(sum(rpes) / len(rpes), 1) if rpes else None

        checks = [
            ("planned_item_count", planned_items),
            ("completed_item_count", completed_items),
            ("planned_working_sets", planned_sets),
            ("completed_working_sets", completed_sets),
        ]
        for field, expected in checks:
            if as_int(s.get(field)) != expected:
                v.error("AGG001", SESSION_SHEET, n, sid, field, "Không khớp dữ liệu item", f"stored={s.get(field)}, expected={expected}")
        if not approx_equal(s.get("completion_pct"), pct(completed_items, planned_items)):
            v.error("AGG002", SESSION_SHEET, n, sid, "completion_pct", "Không khớp aggregate từ item")
        if not approx_equal(s.get("set_completion_pct"), pct(completed_sets, planned_sets)):
            v.error("AGG003", SESSION_SHEET, n, sid, "set_completion_pct", "Không khớp aggregate từ item")
        expected_pain = "Yes" if pain_yes else "No"
        if s.get("pain_reported") != expected_pain:
            v.error("AGG004", SESSION_SHEET, n, sid, "pain_reported", "Không khớp pain của item", f"stored={s.get('pain_reported')}, expected={expected_pain}")
        if mean_rpe is None:
            if not is_blank(s.get("session_rpe")):
                v.error("AGG005", SESSION_SHEET, n, sid, "session_rpe", "Không có actual_rpe item nhưng session_rpe lại có giá trị")
        elif not approx_equal(s.get("session_rpe"), mean_rpe, tol=0.11):
            v.warning("AGG006", SESSION_SHEET, n, sid, "session_rpe", "session_rpe lệch mean actual_rpe của item", f"stored={s.get('session_rpe')}, expected≈{mean_rpe}")

    # Summary must be unique by plan and correspond to a session.
    check_unique_and_id_format(v, summaries, SUMMARY_SHEET, "summary_id", "SUM")
    sessions_by_plan = {str(s.get("plan_id")): s for s in sessions}
    seen_plan: set[str] = set()
    for r in summaries:
        n, sumid = r["__row__"], str(r.get("summary_id") or "")
        pid = str(r.get("plan_id") or "")
        if pid in seen_plan:
            v.error("SUM004", SUMMARY_SHEET, n, sumid, "plan_id", "Một plan có nhiều summary đại diện", pid)
        seen_plan.add(pid)
        s = sessions_by_plan.get(pid)
        if s is None:
            v.error("SUM005", SUMMARY_SHEET, n, sumid, "plan_id", "Summary không có session tương ứng", pid)
            continue
        mappings = [
            ("user_id", "user_id"), ("session_status", "completion_status"),
            ("session_completion_pct", "completion_pct"), ("set_completion_pct", "set_completion_pct"),
            ("session_rpe", "session_rpe"), ("fatigue_after", "fatigue_after"),
            ("pain_reported", "pain_reported"), ("recovery_flag", "recovery_flag"),
        ]
        for sf, sessf in mappings:
            a, b = r.get(sf), s.get(sessf)
            if sf in {"session_completion_pct", "set_completion_pct", "session_rpe", "fatigue_after"}:
                if is_blank(a) and is_blank(b):
                    continue
                if not approx_equal(a, b):
                    v.error("SUM006", SUMMARY_SHEET, n, sumid, sf, f"Không khớp session.{sessf}", f"summary={a}, session={b}")
            elif a != b:
                v.error("SUM006", SUMMARY_SHEET, n, sumid, sf, f"Không khớp session.{sessf}", f"summary={a}, session={b}")

        if r.get("progression_recommendation") not in PROGRESSION:
            v.error("SUM007", SUMMARY_SHEET, n, sumid, "progression_recommendation", f"Giá trị ngoài taxonomy {sorted(PROGRESSION)}", r.get("progression_recommendation"))
        if r.get("pain_reported") == "Yes" and r.get("progression_recommendation") != "REVIEW_BEFORE_PROGRESSION":
            v.warning("SUM008", SUMMARY_SHEET, n, sumid, "progression_recommendation", "Có pain thì nên REVIEW_BEFORE_PROGRESSION", r.get("progression_recommendation"))


def load_master_index(path: Path | None, sheet: str, key: str) -> dict[str, dict]:
    if path is None:
        return {}
    rows = load_rows(path, sheet)
    return {str(r.get(key)): r for r in rows if not is_blank(r.get(key))}


def validate_external_references(v: Validator, sessions: list[dict], items: list[dict],
                                 user_path: Path | None, plan_path: Path | None,
                                 exercise_path: Path | None) -> None:
    users = load_master_index(user_path, "User_Profile", "user_id") if user_path else {}
    plans = load_master_index(plan_path, "Workout_Plan", "plan_id") if plan_path else {}
    plan_items = load_master_index(plan_path, "Workout_Plan_Items", "plan_item_id") if plan_path else {}
    exercises = load_master_index(exercise_path, "gym_exercise_dataset", "exercise_id") if exercise_path else {}

    if user_path:
        for s in sessions:
            uid, sid = str(s.get("user_id") or ""), str(s.get("history_session_id") or "")
            if uid not in users:
                v.error("FK001", SESSION_SHEET, s["__row__"], sid, "user_id", "user_id không tồn tại trong User Master", uid)
            else:
                u = users[uid]
                # Snapshots should be plausible relative to current profile, but may legitimately drift over time.
                current_weight = as_float(u.get("weight_kg"))
                snap_weight = as_float(s.get("body_weight_kg_snapshot"))
                if current_weight is not None and snap_weight is not None and abs(current_weight - snap_weight) > max(10, current_weight * 0.15):
                    v.warning("FK002", SESSION_SHEET, s["__row__"], sid, "body_weight_kg_snapshot", "Snapshot lệch >15% hoặc >10kg so với User Master", f"history={snap_weight}, user={current_weight}")

    if plan_path:
        for s in sessions:
            pid, uid, sid = str(s.get("plan_id") or ""), str(s.get("user_id") or ""), str(s.get("history_session_id") or "")
            p = plans.get(pid)
            if p is None:
                v.error("FK003", SESSION_SHEET, s["__row__"], sid, "plan_id", "plan_id không tồn tại trong Workout Plan Master", pid)
            elif str(p.get("user_id")) != uid:
                v.error("FK004", SESSION_SHEET, s["__row__"], sid, "user_id/plan_id", "Plan không thuộc user này", f"history_user={uid}, plan_user={p.get('user_id')}")

        for r in items:
            iid = str(r.get("history_item_id") or "")
            piid = str(r.get("plan_item_id") or "")
            source = plan_items.get(piid)
            if source is None:
                v.error("FK005", ITEM_SHEET, r["__row__"], iid, "plan_item_id", "Không tồn tại trong Workout_Plan_Items", piid)
                continue
            exact_fields = ["plan_id", "exercise_id", "exercise_name_snapshot", "exercise_order", "planned_sets", "planned_rep_min", "planned_rep_max", "planned_rest_seconds"]
            for field in exact_fields:
                source_value = source.get(field)
                # Some generated XLSX files store formula-backed source fields without
                # cached values. In data_only mode openpyxl then returns None.
                # Do not create a false mismatch when the source value is unavailable.
                if not is_blank(source_value) and r.get(field) != source_value:
                    v.error("FK006", ITEM_SHEET, r["__row__"], iid, field, "History prescription không khớp plan item nguồn", f"history={r.get(field)}, plan={source_value}")
            source_target = source.get("target_intensity")
            if not is_blank(source_target) and not approx_equal(r.get("planned_target_rpe"), source_target):
                v.error("FK007", ITEM_SHEET, r["__row__"], iid, "planned_target_rpe", "Không khớp target_intensity trong plan item", f"history={r.get('planned_target_rpe')}, plan={source_target}")

    if exercise_path:
        for r in items:
            iid, exid = str(r.get("history_item_id") or ""), str(r.get("exercise_id") or "")
            ex = exercises.get(exid)
            if ex is None:
                v.error("FK008", ITEM_SHEET, r["__row__"], iid, "exercise_id", "exercise_id không tồn tại trong Exercise Master", exid)
            elif str(r.get("exercise_name_snapshot")) != str(ex.get("exercise_name")):
                v.warning("FK009", ITEM_SHEET, r["__row__"], iid, "exercise_name_snapshot", "Tên snapshot khác Exercise Master hiện tại; có thể do rename/versioning", f"history={r.get('exercise_name_snapshot')}, master={ex.get('exercise_name')}")


def write_report(path: Path, issues: list[Issue], counts: dict[str, int]) -> None:
    sev = Counter(i.severity for i in issues)
    rule_counts = Counter(i.rule_id for i in issues)
    lines = [
        "WORKOUT HISTORY VALIDATION REPORT",
        "=" * 72,
        f"Sessions checked : {counts.get('sessions', 0)}",
        f"Items checked    : {counts.get('items', 0)}",
        f"Summaries checked: {counts.get('summaries', 0)}",
        "",
        f"ERROR   : {sev.get('ERROR', 0)}",
        f"WARNING : {sev.get('WARNING', 0)}",
        f"INFO    : {sev.get('INFO', 0)}",
        "",
        "RULE COUNTS",
        "-" * 72,
    ]
    for rule, count in sorted(rule_counts.items()):
        lines.append(f"{rule}: {count}")
    lines += ["", "DETAILS", "-" * 72]
    if not issues:
        lines.append("Không phát hiện vấn đề.")
    else:
        for idx, issue in enumerate(issues, 1):
            lines.append(f"{idx}. {issue.render()}")
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate workout_history_master.xlsx")
    parser.add_argument("history", nargs="?", default="workout_history_master.xlsx", help="Workout history workbook")
    parser.add_argument("--user-master", help="Optional user_master.xlsx for FK validation")
    parser.add_argument("--plan-master", help="Optional workout_plan_master.xlsx for FK/prescription validation")
    parser.add_argument("--exercise-master", help="Optional exercise_master.xlsx for exercise FK validation")
    parser.add_argument("--report", default="workout_history_validation_report.txt", help="Output report path")
    parser.add_argument("--fail-on-warning", action="store_true", help="Exit code 2 when warnings exist")
    args = parser.parse_args()

    history_path = Path(args.history)
    if not history_path.exists():
        print(f"[FATAL] Không tìm thấy file: {history_path}")
        return 3

    v = Validator()
    if not validate_schema(v, history_path):
        write_report(Path(args.report), v.issues, {})
        print(f"Validation dừng vì lỗi schema. Report: {args.report}")
        return 1

    sessions = load_rows(history_path, SESSION_SHEET)
    items = load_rows(history_path, ITEM_SHEET)
    summaries = load_rows(history_path, SUMMARY_SHEET)

    validate_sessions(v, sessions)
    validate_items(v, items)
    validate_internal_links(v, sessions, items, summaries)
    validate_external_references(
        v, sessions, items,
        Path(args.user_master) if args.user_master else None,
        Path(args.plan_master) if args.plan_master else None,
        Path(args.exercise_master) if args.exercise_master else None,
    )

    report_path = Path(args.report)
    write_report(report_path, v.issues, {
        "sessions": len(sessions), "items": len(items), "summaries": len(summaries)
    })

    sev = Counter(i.severity for i in v.issues)
    print("=" * 72)
    print("WORKOUT HISTORY VALIDATION")
    print("=" * 72)
    print(f"Sessions : {len(sessions)}")
    print(f"Items    : {len(items)}")
    print(f"Summaries: {len(summaries)}")
    print(f"ERROR    : {sev.get('ERROR', 0)}")
    print(f"WARNING  : {sev.get('WARNING', 0)}")
    print(f"INFO     : {sev.get('INFO', 0)}")
    print(f"Report   : {report_path}")

    if sev.get("ERROR", 0) > 0:
        return 1
    if args.fail_on_warning and sev.get("WARNING", 0) > 0:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
